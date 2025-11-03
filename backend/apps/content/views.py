import csv
from io import TextIOWrapper
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Qualification, Paper, Module, Chapter, Topic, Question
from .serializers import (
    QualificationSerializer,
    PaperSerializer,
    ModuleSerializer,
    ChapterSerializer,
    TopicSerializer,
    QuestionSerializer,
)
from apps.accounts.permissions import IsAdmin, IsTeacher


class QualificationViewSet(viewsets.ModelViewSet):
    queryset = Qualification.objects.all().order_by("name")
    serializer_class = QualificationSerializer
    permission_classes = [permissions.IsAuthenticated]


class PaperViewSet(viewsets.ModelViewSet):
    queryset = Paper.objects.all().order_by("name")
    serializer_class = PaperSerializer
    permission_classes = [permissions.IsAuthenticated]


class ModuleViewSet(viewsets.ModelViewSet):
    queryset = Module.objects.all().order_by("name")
    serializer_class = ModuleSerializer
    permission_classes = [permissions.IsAuthenticated]


class ChapterViewSet(viewsets.ModelViewSet):
    queryset = Chapter.objects.all().order_by("name")
    serializer_class = ChapterSerializer
    permission_classes = [permissions.IsAuthenticated]


class TopicViewSet(viewsets.ModelViewSet):
    queryset = Topic.objects.all().order_by("name")
    serializer_class = TopicSerializer
    permission_classes = [permissions.IsAuthenticated]


class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.filter(is_active=True).order_by("-created_at")
    serializer_class = QuestionSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=["post"], permission_classes=[IsAdmin | IsTeacher])
    def bulk_upload(self, request):
        """
        Expect multipart/form-data with file=<csv|xlsx>
        Columns required:
        qualification,paper,module,chapter,topic,qtype,difficulty,skill_type,text,explanation,marks,professional_marks,choices_json,correct_json
        """
        import os
        import json
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "file is required"}, status=400)

        created = 0
        name = getattr(file, 'name', '')
        ext = os.path.splitext(name)[1].lower()
        rows = []
        if ext in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename=file, read_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for r in ws.iter_rows(min_row=2, values_only=True):
                    row = { headers[i]: (str(v).strip() if isinstance(v, str) else v) for i, v in enumerate(r) }
                    rows.append(row)
            except Exception as e:
                return Response({"detail": f"Excel parse failed: {e}"}, status=400)
        else:
            reader = csv.DictReader(TextIOWrapper(file.file, encoding="utf-8"))
            rows = list(reader)

        for row in rows:
            try:
                q_name = str(row.get("qualification", "")).strip()
                p_name = str(row.get("paper", "")).strip()
                m_name = str(row.get("module", "")).strip()
                c_name = str(row.get("chapter", "")).strip()
                t_name = str(row.get("topic", "")).strip()
                if not (q_name and p_name and m_name and c_name and t_name):
                    continue
                qual, _ = Qualification.objects.get_or_create(name=q_name)
                paper, _ = Paper.objects.get_or_create(qualification=qual, name=p_name)
                module, _ = Module.objects.get_or_create(paper=paper, name=m_name)
                chapter, _ = Chapter.objects.get_or_create(module=module, name=c_name)
                topic, _ = Topic.objects.get_or_create(chapter=chapter, name=t_name)

                Question.objects.create(
                    qualification=qual,
                    paper=paper,
                    module=module,
                    chapter=chapter,
                    topic=topic,
                    qtype=str(row.get("qtype", "MCQ_SINGLE")).strip(),
                    difficulty=str(row.get("difficulty", "MEDIUM")).strip(),
                    skill_type=str(row.get("skill_type", "KNOWLEDGE")).strip(),
                    text=str(row.get("text", "")).strip(),
                    explanation=str(row.get("explanation", "")).strip(),
                    marks=float(row.get("marks", 1) or 1),
                    professional_marks=float(row.get("professional_marks", 0) or 0),
                    choices=(json.loads(row.get("choices_json") or "[]") if isinstance(row.get("choices_json"), str) else (row.get("choices_json") or [])),
                    correct_answer=(json.loads(row.get("correct_json") or "{}") if isinstance(row.get("correct_json"), str) else (row.get("correct_json") or {})),
                )
                created += 1
            except Exception:
                continue
        return Response({"created": created})
